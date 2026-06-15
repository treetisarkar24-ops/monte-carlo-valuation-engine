#!/usr/bin/env python3
"""
excel_loader.py — packaging / usability layer for the Monte Carlo Valuation Engine.

Reads a filled-in copy of MC_Input_Template.xlsx, validates the inputs, and
converts them into the DCFInputs JSON the engine consumes.

This module sits in FRONT of the engine. It does NOT import-and-alter or modify
any DCF, Monte Carlo, shock, or convergence logic — it only reads a spreadsheet
and produces the same JSON shape the engine already accepts. The validation here
is an input gate for the spreadsheet, not the deferred engine-level Step 7
(which would live inside DCFInputs itself).

Usage:
    python3 excel_loader.py filled_template.xlsx                 # print JSON to stdout
    python3 excel_loader.py filled_template.xlsx -o inputs.json  # write JSON to a file
    python3 excel_loader.py filled_template.xlsx --run           # validate, then run a quick valuation
"""
from __future__ import annotations

# --- engine path bootstrap (auto-added during reorg; frozen engine lives in ../engine) ---
import os as _os, sys as _sys
_sys.path.insert(0, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "..", "engine"))
# --- end bootstrap ---

import argparse
import json
import sys

from openpyxl import load_workbook

# Field order mirrors DCFInputs. Kept as a local constant so the loader stays
# decoupled from the engine for the parse/validate steps.
SCALARS = [
    "starting_revenue", "net_debt", "shares_outstanding", "forecast_years",
    "tax_rate", "terminal_growth", "risk_free_rate", "equity_risk_premium",
    "beta", "cost_of_debt", "debt_to_total_capital",
]
TRAJECTORIES = [
    "revenue_growth", "operating_margin", "capex_pct_revenue",
    "da_pct_revenue", "nwc_pct_revenue",
]
INT_FIELDS = {"forecast_years"}


class TemplateError(ValueError):
    """Raised when the filled template cannot be turned into valid engine inputs."""


def _to_number(value):
    """Coerce a cell value to float, rejecting booleans (bool is an int subclass)."""
    if isinstance(value, bool):
        raise ValueError("boolean where a number was expected")
    return float(value)


def read_template(path, sheet="Inputs"):
    """Parse the raw scalar map and trajectory rows out of a filled template.

    Returns (raw_scalars: dict[str, cell], traj_rows: list[tuple]).
    No validation here — that happens in build_inputs().
    """
    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise TemplateError(
            f"sheet '{sheet}' not found in {path}. Sheets present: {wb.sheetnames}"
        )
    ws = wb[sheet]
    grid = [list(r) + [None] * (6 - len(r)) for r in ws.iter_rows(values_only=True)]

    # Block A — scalars: column A (idx 0) carries the field name, column E
    # (idx 4) carries the user's value.
    raw_scalars = {}
    for row in grid:
        key = row[0]
        if isinstance(key, str) and key.strip() in SCALARS:
            raw_scalars[key.strip()] = row[4]

    # Block B — trajectories: collect year rows that appear AFTER the
    # "YOUR VALUES ->" marker so the MSFT example rows above it are ignored.
    start = None
    for i, row in enumerate(grid):
        if isinstance(row[0], str) and "YOUR VALUES" in row[0].upper():
            start = i + 1
            break

    traj_rows = []
    if start is not None:
        for row in grid[start:]:
            year_cell = row[0]
            is_year = isinstance(year_cell, (int, float)) and not isinstance(year_cell, bool)
            if not is_year:
                break  # reached the footnote / end of the block
            values = tuple(row[1:6])
            if all(v is None for v in values):
                continue  # an unused year-row (user filled fewer than provided)
            traj_rows.append(values)

    return raw_scalars, traj_rows


def build_inputs(raw_scalars, traj_rows):
    """Validate parsed cells and assemble the DCFInputs JSON dict.

    Returns (data: dict, warnings: list[str]). Raises TemplateError listing
    every problem found, so the user fixes them in one pass rather than one at
    a time.
    """
    errors = []
    warnings = []
    data = {}

    # --- scalars ---
    for field in SCALARS:
        raw = raw_scalars.get(field)
        if raw is None or (isinstance(raw, str) and raw.strip() == ""):
            errors.append(f"missing value for '{field}' (Block A, YOUR VALUE column)")
            continue
        try:
            num = _to_number(raw)
        except (TypeError, ValueError):
            errors.append(f"'{field}' must be a number, got {raw!r}")
            continue
        if field in INT_FIELDS:
            if num <= 0 or num != int(num):
                errors.append(f"'{field}' must be a positive whole number, got {raw!r}")
                continue
            num = int(num)
        data[field] = num

    forecast_years = data.get("forecast_years")

    # --- trajectories ---
    if not traj_rows:
        errors.append(
            "no year-by-year trajectory rows found under 'YOUR VALUES ->' in Block B"
        )
    else:
        cols = {name: [] for name in TRAJECTORIES}
        for row_idx, values in enumerate(traj_rows, start=1):
            for j, name in enumerate(TRAJECTORIES):
                raw = values[j]
                if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                    errors.append(f"trajectory '{name}' missing a value in year row {row_idx}")
                    cols[name].append(None)
                    continue
                try:
                    cols[name].append(_to_number(raw))
                except (TypeError, ValueError):
                    errors.append(f"trajectory '{name}' year {row_idx} must be a number, got {raw!r}")
                    cols[name].append(None)
        for name in TRAJECTORIES:
            data[name] = cols[name]
        if forecast_years is not None and len(traj_rows) != forecast_years:
            errors.append(
                f"filled {len(traj_rows)} trajectory year-row(s) but forecast_years = "
                f"{forecast_years}; the two must match"
            )

    # --- range / decimal-vs-percent sanity ---
    def warn_if_outside(field, lo, hi):
        v = data.get(field)
        if isinstance(v, (int, float)) and not (lo <= v <= hi):
            warnings.append(
                f"'{field}' = {v} is outside the usual {lo}..{hi} for a decimal rate "
                f"— did you enter a percent (e.g. 19 instead of 0.19)?"
            )

    for field in ["terminal_growth", "risk_free_rate", "equity_risk_premium", "cost_of_debt"]:
        warn_if_outside(field, 0.0, 1.0)

    # hard bounds that would otherwise break the math
    tax = data.get("tax_rate")
    if isinstance(tax, (int, float)) and not (0.0 <= tax < 1.0):
        errors.append(f"tax_rate must be in [0, 1), got {tax}")
    dtc = data.get("debt_to_total_capital")
    if isinstance(dtc, (int, float)) and not (0.0 <= dtc <= 1.0):
        errors.append(f"debt_to_total_capital must be in [0, 1], got {dtc}")

    if errors:
        raise TemplateError("Template has problems:\n  - " + "\n  - ".join(errors))

    # --- terminal_growth < WACC ---
    # Uses the engine's own WACC computation (read-only call; nothing modified).
    try:
        from dcf import DCFInputs, compute_wacc
        wacc = compute_wacc(DCFInputs(**data))
        if data["terminal_growth"] >= wacc:
            raise TemplateError(
                f"terminal_growth ({data['terminal_growth']:.4f}) must be LESS than WACC "
                f"({wacc:.4f}); otherwise the terminal-value formula diverges."
            )
    except ImportError:
        warnings.append(
            "could not import the engine to verify terminal_growth < WACC "
            "(run from the MC folder to enable this check)"
        )

    return data, warnings


def _is_typed_template(grid):
    """Detect the typed layout: a header row carrying an 'engine_field' column."""
    for row in grid[:12]:
        for cell in row:
            if isinstance(cell, str) and cell.strip().lower() == "engine_field":
                return True
    return False


def read_template_typed(path, sheet="Inputs"):
    """Parse the typed template (Line item | engine_field | Unit | Example | values...).

    Keys every row on its `engine_field` tag in column B (idx 1). Scalars take the
    single value in the 'YOUR VALUE' column (idx 4); trajectories take the per-year
    values running from idx 4 across the Year columns. Returns the SAME
    (raw_scalars, traj_rows) shape as read_template(), so build_inputs() is reused
    verbatim — no validation logic is duplicated.
    """
    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise TemplateError(
            f"sheet '{sheet}' not found in {path}. Sheets present: {wb.sheetnames}"
        )
    ws = wb[sheet]
    width = max(5, ws.max_column)
    grid = [list(r) + [None] * (width - len(r)) for r in ws.iter_rows(values_only=True)]

    raw_scalars = {}
    traj_cells = {}  # engine_field -> list of per-year cells (idx 4 onward)
    for row in grid:
        tag = row[1]
        if not isinstance(tag, str):
            continue
        tag = tag.strip()
        if tag in SCALARS:
            raw_scalars[tag] = row[4]
        elif tag in TRAJECTORIES:
            traj_cells[tag] = list(row[4:])

    # Determine the number of year columns to read. Prefer forecast_years; if it
    # is missing/unusable, fall back to the longest filled run so build_inputs()
    # can still surface the real problem rather than crashing here.
    fy_raw = raw_scalars.get("forecast_years")
    n = None
    if isinstance(fy_raw, (int, float)) and not isinstance(fy_raw, bool) and fy_raw == int(fy_raw) and fy_raw > 0:
        n = int(fy_raw)
    if n is None:
        n = 0
        for cells in traj_cells.values():
            filled = [i for i, v in enumerate(cells)
                      if not (v is None or (isinstance(v, str) and v.strip() == ""))]
            if filled:
                n = max(n, filled[-1] + 1)
        n = n or 1

    def cell_for(field, year_idx):
        cells = traj_cells.get(field, [])
        return cells[year_idx] if year_idx < len(cells) else None

    traj_rows = []
    for yr in range(n):
        traj_rows.append(tuple(cell_for(name, yr) for name in TRAJECTORIES))

    return raw_scalars, traj_rows


def excel_to_inputs(path):
    """Convenience: parse + validate a filled template, returning (data, warnings).

    Auto-detects the layout: the typed template (with an 'engine_field' column) or
    the original block template. Both feed the identical build_inputs() validator.
    """
    wb = load_workbook(path, data_only=True)
    sheet = "Inputs" if "Inputs" in wb.sheetnames else wb.sheetnames[0]
    grid = [list(r) for r in wb[sheet].iter_rows(values_only=True)]
    if _is_typed_template(grid):
        raw_scalars, traj_rows = read_template_typed(path, sheet=sheet)
    else:
        raw_scalars, traj_rows = read_template(path, sheet=sheet)
    return build_inputs(raw_scalars, traj_rows)


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Convert a filled MC input template (.xlsx) into engine JSON."
    )
    parser.add_argument("xlsx", help="filled copy of MC_Input_Template.xlsx")
    parser.add_argument("-o", "--out", help="write JSON here (default: print to stdout)")
    parser.add_argument("--run", action="store_true", help="after loading, run a quick valuation")
    args = parser.parse_args(argv)

    try:
        data, warnings = excel_to_inputs(args.xlsx)
    except (TemplateError, FileNotFoundError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    for w in warnings:
        print(f"WARNING: {w}", file=sys.stderr)

    text = json.dumps(data, indent=2)
    if args.out:
        with open(args.out, "w") as f:
            f.write(text + "\n")
        print(f"wrote {args.out}")
    else:
        print(text)

    if args.run:
        from dcf import DCFInputs, run_dcf, compute_wacc
        from mc_config import MCConfig
        from mc_engine import run_monte_carlo, summarize
        inp = DCFInputs(**data)
        print(f"\nWACC          : {compute_wacc(inp) * 100:.2f}%")
        print(f"Deterministic : {run_dcf(inp):.2f} / share")
        res = run_monte_carlo(inp, MCConfig(n_simulations=2000, random_seed=42))
        summary = summarize(res)
        print(
            f"MC median     : {summary.median:.2f}   "
            f"(P5 {summary.percentiles[5]:.2f} / P95 {summary.percentiles[95]:.2f}, n={summary.n})"
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
