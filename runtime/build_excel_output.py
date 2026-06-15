#!/usr/bin/env python3
"""
build_excel_output.py — supporting deliverable: a results workbook for power users.

Second tier of the three-tier output set (PDF report / Excel workbook / Markdown
diagnostics). Reads the SAME results.json + meta.json + fixture as the other two,
and lays the engine's outputs out across labelled sheets. Presentation layer only;
the engine is imported read-only for the year-by-year projection.

Usage:
    python3 build_excel_output.py amazon_results.json amazon_meta.json -o Amazon_Output.xlsx
"""
from __future__ import annotations

# --- engine path bootstrap (auto-added during reorg; frozen engine lives in ../engine) ---
import os as _os, sys as _sys
_sys.path.insert(0, _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "..", "engine"))
# --- end bootstrap ---

import argparse
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3A5F"
LIGHT = "EAF0F6"
HDRF = PatternFill("solid", fgColor=NAVY)
LITF = PatternFill("solid", fgColor=LIGHT)
WHITE_B = Font(bold=True, color="FFFFFF", name="Arial")
BOLD = Font(bold=True, name="Arial")
ARIAL = Font(name="Arial")
BLUE = Font(name="Arial", color="0000FF")  # hardcoded inputs (industry convention)
NAVYF = Font(name="Arial", bold=True, color=NAVY)
WRAP = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="C9D4E0")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

MONEY = '$#,##0.00;($#,##0.00);-'
MONEY0 = '$#,##0;($#,##0);-'
PCT = '0.0%'
PCT2 = '0.00%'


def hdr(ws, row, labels, widths=None):
    for j, lab in enumerate(labels, start=1):
        c = ws.cell(row=row, column=j, value=lab)
        c.font = WHITE_B; c.fill = HDRF; c.border = BORD
        c.alignment = Alignment(vertical="center")
    if widths:
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w


def title(ws, text):
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(bold=True, size=14, color=NAVY, name="Arial")


def zebra(ws, r):
    if r % 2 == 0:
        for c in ws[r]:
            if c.value is not None:
                c.fill = LITF


def year_tables(meta):
    fixture = meta.get("fixture")
    if not fixture:
        return None, None, None
    try:
        from dcf import (DCFInputs, project_revenue, project_ebit, project_nopat,
                         project_da, project_capex, project_dnwc, project_fcf)
        path = fixture
        if not os.path.exists(path) and meta.get("__meta_path__"):
            path = os.path.join(os.path.dirname(meta["__meta_path__"]), fixture)
        inp = DCFInputs(**json.load(open(path)))
    except Exception:
        return None, None, None
    base = meta.get("base_year")
    fy = inp.forecast_years
    rev = project_revenue(inp); ebit = project_ebit(rev, inp); nopat = project_nopat(ebit, inp)
    da = project_da(rev, inp); capex = project_capex(rev, inp); dnwc = project_dnwc(rev, inp)
    fcf = project_fcf(nopat, rev, inp)
    label = (lambda i: f"FY{base + 1 + i}") if base else (lambda i: f"Year {i+1}")
    rows = []
    for i in range(fy):
        rows.append(dict(year=label(i), g=inp.revenue_growth[i], om=inp.operating_margin[i],
                         cx=inp.capex_pct_revenue[i], da=inp.da_pct_revenue[i], nwc=inp.nwc_pct_revenue[i],
                         rev=rev[i], ebit=ebit[i], nopat=nopat[i], dav=da[i], capexv=capex[i], fcf=fcf[i]))
    base_label = f"FY{base}" if base else "Year 0"
    return rows, base_label, inp.starting_revenue


def build(results, meta, out_path):
    central = results["central"]; cont = results["cont"]; shock = results.get("shock", {})
    wb = Workbook()

    # ---------- Summary ----------
    ws = wb.active; ws.title = "Summary"
    title(ws, f"{meta['company']} ({meta['ticker']}) — Valuation Summary")
    rows = [
        ("Market price", meta["market_price"], MONEY, BLUE),
        ("Market date", meta.get("market_date", ""), None, ARIAL),
        ("Central DCF / share", central["value"], MONEY, NAVYF),
        ("WACC", central["wacc"], PCT2, ARIAL),
        ("Terminal growth", meta["terminal_growth"], PCT2, BLUE),
        ("Continuous MC median", cont["production"]["median"], MONEY, ARIAL),
        ("Continuous MC mean", cont["production"]["mean"], MONEY, ARIAL),
        ("Fraction negative (cont)", cont["frac_negative_pct"] / 100.0, PCT, ARIAL),
    ]
    if shock:
        rows += [
            ("Shocked MC median", shock["production"]["median"], MONEY, ARIAL),
            ("Shock-free paths", shock["shock_free_pct"] / 100.0, PCT, ARIAL),
            ("Fraction negative (shock)", shock["frac_negative_pct"] / 100.0, PCT, ARIAL),
        ]
    r = 3
    hdr(ws, r, ["Metric", "Value"], widths=[30, 18]); r += 1
    for label, val, fmt, fnt in rows:
        ws.cell(row=r, column=1, value=label).font = ARIAL
        c = ws.cell(row=r, column=2, value=val); c.font = fnt
        if fmt:
            c.number_format = fmt
        for cc in ws[r][:2]:
            cc.border = BORD
        zebra(ws, r); r += 1
    # one derived formula: terminal spread = WACC - g
    ws.cell(row=r, column=1, value="Terminal spread (WACC − g)").font = ARIAL
    wacc_row = 7  # data rows start at 4: Market price(4) Market date(5) Central(6) WACC(7) Terminal growth(8)
    g_row = 8
    sc = ws.cell(row=r, column=2, value=f"=B{wacc_row}-B{g_row}")
    sc.number_format = PCT2; sc.font = ARIAL
    for cc in ws[r][:2]:
        cc.border = BORD
    r += 2
    ws.cell(row=r, column=1,
            value="Source: Monte Carlo Valuation Engine run (seed 42). Engine applied "
                  "as-is; no DCF/MC/shock/convergence logic modified.").font = Font(italic=True, size=8, name="Arial")

    # ---------- Assumptions ----------
    ws = wb.create_sheet("Assumptions")
    title(ws, "DCFInputs Assumptions")
    hdr(ws, 3, ["Input", "Value", "Rationale"], widths=[24, 14, 70])
    r = 4
    for inp, val, why in meta["assumptions"]:
        ws.cell(row=r, column=1, value=inp).font = ARIAL
        ws.cell(row=r, column=2, value=val).font = BLUE
        wc = ws.cell(row=r, column=3, value=why); wc.font = ARIAL; wc.alignment = WRAP
        for cc in ws[r][:3]:
            cc.border = BORD
        zebra(ws, r); r += 1

    # ---------- Deterministic ----------
    ws = wb.create_sheet("Deterministic")
    title(ws, "Deterministic Central Case")
    yr_rows, base_label, start_rev = year_tables(meta)
    r = 3
    if yr_rows is not None:
        hdr(ws, r, ["Year", "Rev growth", "Op margin", "Capex %", "D&A %", "ΔNWC %"],
            widths=[12, 12, 12, 12, 12, 12]); r += 1
        for d in yr_rows:
            ws.cell(row=r, column=1, value=d["year"]).font = ARIAL
            for j, key, fmt in [(2, "g", PCT), (3, "om", PCT), (4, "cx", PCT), (5, "da", PCT), (6, "nwc", PCT)]:
                c = ws.cell(row=r, column=j, value=d[key]); c.number_format = fmt; c.font = BLUE; c.border = BORD
            ws.cell(row=r, column=1).border = BORD
            zebra(ws, r); r += 1
        r += 2
        hdr(ws, r, ["Year", "Revenue", "EBIT", "NOPAT", "+ D&A", "− Capex", "FCFF"],
            widths=[12, 14, 14, 14, 12, 12, 14]); r += 1
        ws.cell(row=r, column=1, value=base_label).font = ARIAL
        bc = ws.cell(row=r, column=2, value=start_rev); bc.number_format = MONEY0; bc.font = ARIAL
        for cc in ws[r][:7]:
            cc.border = BORD
        r += 1
        for d in yr_rows:
            ws.cell(row=r, column=1, value=d["year"]).font = ARIAL
            for j, key in [(2, "rev"), (3, "ebit"), (4, "nopat"), (5, "dav"), (6, "capexv"), (7, "fcf")]:
                v = -d[key] if key == "capexv" else d[key]
                c = ws.cell(row=r, column=j, value=v); c.number_format = MONEY0; c.font = ARIAL; c.border = BORD
            ws.cell(row=r, column=1).border = BORD
            zebra(ws, r); r += 1
        ws.cell(row=r + 1, column=1, value="All currency figures in fixture units (USD millions).").font = Font(italic=True, size=8, name="Arial")

    # ---------- Distributions ----------
    def dist_sheet(name, blk, frac_neg):
        ws = wb.create_sheet(name)
        title(ws, name)
        pair = blk["pair"]; prod = blk["production"]
        hdr(ws, 3, ["Statistic", "Value"], widths=[26, 16])
        stats = [
            ("z (paths)", pair["z_star"], "0"),
            ("Decision margin", pair["decision_margin_pct"] / 100.0, PCT),
            ("z_pct", pair["z_pct"], "0"),
            ("z_elbow", pair["z_elbow"], "0"),
            ("Mean", prod["mean"], MONEY),
            ("Median", prod["median"], MONEY),
            ("Std dev", prod["std"], MONEY),
            ("Min", prod["min"], MONEY),
            ("Max", prod["max"], MONEY),
            ("Fraction negative", frac_neg / 100.0, PCT),
        ]
        r = 4
        for label, val, fmt in stats:
            ws.cell(row=r, column=1, value=label).font = ARIAL
            c = ws.cell(row=r, column=2, value=val); c.number_format = fmt; c.font = ARIAL
            for cc in ws[r][:2]:
                cc.border = BORD
            zebra(ws, r); r += 1
        r += 1
        hdr(ws, r, ["Percentile", "Value"]); r += 1
        for k in sorted(prod["pct"], key=lambda x: int(x)):
            ws.cell(row=r, column=1, value=f"P{int(k)}").font = ARIAL
            c = ws.cell(row=r, column=2, value=prod["pct"][k]); c.number_format = MONEY; c.font = ARIAL
            for cc in ws[r][:2]:
                cc.border = BORD
            zebra(ws, r); r += 1
        return ws

    dist_sheet("Continuous MC", cont, cont["frac_negative_pct"])
    if shock:
        wss = dist_sheet("Shocked MC", shock, shock["frac_negative_pct"])
        if shock.get("fires_all"):
            r = wss.max_row + 2
            chans = list(shock["fires_all"].keys())
            hdr(wss, r, ["Channel", "Fires (all)", "Fires (worst 5%)"], widths=[16, 14, 16]); r += 1
            for c in chans:
                wss.cell(row=r, column=1, value=c.capitalize()).font = ARIAL
                wss.cell(row=r, column=2, value=shock["fires_all"][c]).font = ARIAL
                wss.cell(row=r, column=3, value=shock["fires_worst5"][c]).font = ARIAL
                for cc in wss[r][:3]:
                    cc.border = BORD
                zebra(wss, r); r += 1

    # ---------- Seeds ----------
    if results.get("seeds"):
        ws = wb.create_sheet("Seed Robustness")
        title(ws, "Seed Robustness")
        hdr(ws, 3, ["Seed", "z* (cont)", "cont margin", "z** (shock)", "shock margin"],
            widths=[10, 12, 14, 12, 14])
        r = 4
        for seed, d in results["seeds"].items():
            ws.cell(row=r, column=1, value=str(seed)).font = ARIAL
            ws.cell(row=r, column=2, value=d.get("cont_z")).font = ARIAL
            if "cont_margin" in d:
                c = ws.cell(row=r, column=3, value=d["cont_margin"] / 100.0); c.number_format = PCT; c.font = ARIAL
            ws.cell(row=r, column=4, value=d.get("shock_z")).font = ARIAL
            if "shock_margin" in d:
                c = ws.cell(row=r, column=5, value=d["shock_margin"] / 100.0); c.number_format = PCT; c.font = ARIAL
            for cc in ws[r][:5]:
                cc.border = BORD
            zebra(ws, r); r += 1

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"
    wb.save(out_path)
    return out_path


def main(argv=None):
    ap = argparse.ArgumentParser(description="Build a results workbook (.xlsx) from a case study.")
    ap.add_argument("results")
    ap.add_argument("meta")
    ap.add_argument("-o", "--out", required=True)
    args = ap.parse_args(argv)
    results = json.load(open(args.results))
    meta = json.load(open(args.meta))
    meta["__meta_path__"] = os.path.abspath(args.meta)
    build(results, meta, args.out)
    print(f"wrote {args.out}")


if __name__ == "__main__":
    raise SystemExit(main())
